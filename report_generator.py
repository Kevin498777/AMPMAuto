# report_generator.py - VERSIÓN SIMPLIFICADA (SOLO EXCEL)
import os
import pandas as pd
from datetime import datetime
import sys
import logging
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

def get_app_data_path():
    """Obtener ruta en AppData/Local para archivos de usuario - MEJORADO"""
    try:
        if os.name == 'nt':  # Windows
            app_data = os.getenv('LOCALAPPDATA')
            if app_data:
                app_path = os.path.join(app_data, 'AMPMAuto')
                os.makedirs(app_path, exist_ok=True)
                return app_path
        # Fallback para Linux/Mac y casos de error
        fallback_path = os.path.join(os.path.expanduser('~'), '.ampmauto')
        os.makedirs(fallback_path, exist_ok=True)
        return fallback_path
    except Exception as e:
        logger.warning(f"No se pudo crear directorio en AppData: {e}")
        # Último fallback: directorio actual
        current_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        reports_dir = os.path.join(current_dir, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        return reports_dir

class ReportGenerator:
    """Generador de reportes para AMPMAuto - VERSIÓN SIMPLIFICADA (SOLO EXCEL)"""
    
    def __init__(self, output_dir: Optional[str] = None):
        # Usar carpeta con permisos de escritura garantizados
        if output_dir is None:
            output_dir = os.path.join(get_app_data_path(), "reports")
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        logger.info(f"📁 Reportes Excel se guardarán en: {self.output_dir}")
    
    def generate_report(self, report_data: Optional[Dict] = None) -> Optional[str]:
        """Genera un reporte en Excel con los resultados de la automatización"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Validar datos de entrada
            if report_data is None:
                logger.warning("No se proporcionaron datos, generando reporte de ejemplo")
                report_data = self._get_sample_data()
            elif not self._validate_report_data(report_data):
                logger.error("Datos de reporte inválidos")
                return None
            
            # Generar reporte Excel
            excel_path = os.path.join(self.output_dir, f"reporte_ampm_{timestamp}.xlsx")
            success = self._generate_excel_report(excel_path, report_data)
            
            if not success:
                logger.error("Falló la generación del reporte Excel")
                return None
            
            logger.info(f"📊 Reporte Excel generado: {excel_path}")
            return excel_path
            
        except Exception as e:
            logger.error(f"❌ Error crítico al generar reporte: {str(e)}")
            return None
    
    def _validate_report_data(self, report_data: Dict) -> bool:
        """Validar estructura de datos del reporte"""
        required_fields = ['total', 'success', 'errors', 'timestamp', 'results']
        for field in required_fields:
            if field not in report_data:
                logger.error(f"Campo requerido faltante: {field}")
                return False
        return True
    
    def _get_sample_data(self) -> Dict:
        """Generar datos de ejemplo para testing"""
        return {
            'total': 15,
            'success': 12,
            'errors': 3,
            'recovered': 1,
            'timestamp': datetime.now(),
            'excel_file': 'archivo_ejemplo.xlsx',
            'results': [
                {'success': True, 'guia_number': '45123456789', 'message': 'Procesado exitosamente', 'processing_time': 12.5},
                {'success': True, 'guia_number': '45123456790', 'message': 'Procesado exitosamente', 'processing_time': 11.2},
                {'success': False, 'guia_number': '45123456791', 'error': 'Guía ya entregada', 'recoverable': True},
                {'success': False, 'guia_number': '45123456792', 'error': 'Error de conexión', 'recoverable': False},
            ]
        }
    
    def _generate_excel_report(self, file_path: str, report_data: Dict) -> bool:
        """Genera un reporte en Excel con formato profesional"""
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Hoja de resumen
                summary_data = {
                    'Métrica': [
                        'Fecha de Generación',
                        'Archivo Procesado', 
                        'Total de Guías',
                        'Guías Exitosas',
                        'Guías Ya Entregadas',
                        'Errores Reales',
                        'Tasa de Éxito'
                    ],
                    'Valor': [
                        report_data['timestamp'].strftime('%d/%m/%Y %H:%M:%S'),
                        report_data.get('excel_file', 'N/A'),
                        report_data['total'],
                        report_data['success'],
                        report_data.get('recovered', 0),
                        report_data['errors'] - report_data.get('recovered', 0),
                        f"{(report_data['success'] / report_data['total'] * 100):.1f}%" if report_data['total'] > 0 else "0%"
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Hoja de resultados detallados
                if report_data['results']:
                    detailed_data = []
                    for result in report_data['results']:
                        row = {
                            'Guía': result.get('guia_number', 'N/A'),
                            'Estado': 'ÉXITO' if result['success'] else ('YA ENTREGADA' if result.get('recoverable') else 'ERROR'),
                            'Tiempo_Procesamiento_Segundos': result.get('processing_time', 'N/A'),
                            'Mensaje_Error': result.get('message', result.get('error', 'N/A')),
                            'Timestamp': result.get('timestamp', 'N/A'),
                            'Recuperable': 'Sí' if result.get('recoverable') else 'No'
                        }
                        detailed_data.append(row)
                    
                    df_detailed = pd.DataFrame(detailed_data)
                    df_detailed.to_excel(writer, sheet_name='Detalle_Guías', index=False)
                
                # ✅ NUEVA HOJA: GUÍAS CON ERROR (solo errores reales, no recuperables)
                if report_data['results']:
                    error_guides = []
                    for result in report_data['results']:
                        # Solo incluir guías que fallaron y NO son recuperables (errores reales)
                        if not result['success'] and not result.get('recoverable', False):
                            error_row = {
                                'Guía': result.get('guia_number', 'N/A'),
                                'Error': result.get('error', 'Error desconocido'),
                                'Timestamp': result.get('timestamp', 'N/A'),
                                'Tipo_Error': self._classify_error(result.get('error', ''))
                            }
                            error_guides.append(error_row)
                    
                    if error_guides:
                        df_errors = pd.DataFrame(error_guides)
                        df_errors.to_excel(writer, sheet_name='Guias_Error', index=False)
                    else:
                        # Crear hoja vacía si no hay errores
                        df_errors = pd.DataFrame(columns=['Guía', 'Error', 'Timestamp', 'Tipo_Error'])
                        df_errors.to_excel(writer, sheet_name='Guias_Error', index=False)
                
                # Obtener workbook para formateo
                workbook = writer.book
                
                # Formatear hoja de resumen
                if 'Resumen' in workbook.sheetnames:
                    worksheet = workbook['Resumen']
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 30
                
                # Formatear hoja de detalle
                if 'Detalle_Guías' in workbook.sheetnames:
                    worksheet = workbook['Detalle_Guías']
                    worksheet.column_dimensions['A'].width = 15
                    worksheet.column_dimensions['B'].width = 12
                    worksheet.column_dimensions['C'].width = 10
                    worksheet.column_dimensions['D'].width = 40
                    worksheet.column_dimensions['E'].width = 20
                    worksheet.column_dimensions['F'].width = 12
                
                # Formatear hoja de guías con error
                if 'Guias_Error' in workbook.sheetnames:
                    worksheet = workbook['Guias_Error']
                    worksheet.column_dimensions['A'].width = 15
                    worksheet.column_dimensions['B'].width = 50
                    worksheet.column_dimensions['C'].width = 20
                    worksheet.column_dimensions['D'].width = 20
                    
                    # Aplicar formato condicional para resaltar errores
                    from openpyxl.styles import PatternFill, Font
                    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    
                    # Aplicar a todas las filas con datos
                    for row in range(2, len(error_guides) + 2):  # +2 porque empieza en 1 y header en 1
                        worksheet[f'A{row}'].fill = red_fill
                        worksheet[f'B{row}'].fill = red_fill
                        worksheet[f'C{row}'].fill = red_fill
                        worksheet[f'D{row}'].fill = red_fill
            
            logger.info(f"✅ Reporte Excel generado: {file_path}")
            logger.info(f"📊 Hojas incluidas: Resumen, Detalle_Guías, Guias_Error")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error al generar Excel: {str(e)}")
            return False
    
    def _classify_error(self, error_message: str) -> str:
        """Clasificar el tipo de error para mejor análisis"""
        error_lower = error_message.lower()
        
        if any(word in error_lower for word in ['timeout', 'timed out', 'time out']):
            return 'TIMEOUT'
        elif any(word in error_lower for word in ['conexión', 'connection', 'network']):
            return 'ERROR_CONEXION'
        elif any(word in error_lower for word in ['elemento', 'element', 'not found', 'no encontrado']):
            return 'ELEMENTO_NO_ENCONTRADO'
        elif any(word in error_lower for word in ['navegador', 'browser', 'chrome']):
            return 'ERROR_NAVEGADOR'
        elif any(word in error_lower for word in ['login', 'credenciales', 'password']):
            return 'ERROR_LOGIN'
        elif any(word in error_lower for word in ['captcha', 'verificación']):
            return 'CAPTCHA'
        else:
            return 'OTRO_ERROR'
    
    def _calculate_avg_time(self, results: List[Dict]) -> str:
        """Calcular tiempo promedio de procesamiento"""
        valid_times = [r.get('processing_time', 0) for r in results if r.get('processing_time')]
        if valid_times:
            avg = sum(valid_times) / len(valid_times)
            return f"{avg:.2f}s"
        return "N/A"
    
    def _calculate_guides_per_minute(self, results: List[Dict]) -> str:
        """Calcular guías procesadas por minuto"""
        total_time = sum(r.get('processing_time', 0) for r in results if r.get('processing_time'))
        if total_time > 0:
            gpm = len(results) / (total_time / 60)
            return f"{gpm:.1f}"
        return "N/A"

# Funciones de conveniencia para uso rápido
def generate_quick_report(success_count: int, error_count: int, total_count: int, recovered_count: int = 0) -> Optional[str]:
    """Genera un reporte rápido con datos básicos"""
    generator = ReportGenerator()
    report_data = {
        'total': total_count,
        'success': success_count,
        'errors': error_count,
        'recovered': recovered_count,
        'timestamp': datetime.now(),
        'results': []
    }
    return generator.generate_report(report_data)

def generate_detailed_report(results: List[Dict], excel_file: str = "") -> Optional[str]:
    """Genera un reporte detallado con resultados específicos"""
    generator = ReportGenerator()
    
    success_count = sum(1 for r in results if r.get('success'))
    recovered_count = sum(1 for r in results if r.get('recoverable') and not r.get('success'))
    error_count = len(results) - success_count - recovered_count
    
    report_data = {
        'total': len(results),
        'success': success_count,
        'errors': error_count + recovered_count,  # Total de no-éxitos
        'recovered': recovered_count,
        'timestamp': datetime.now(),
        'excel_file': excel_file,
        'results': results
    }
    
    return generator.generate_report(report_data)

# Ejemplo de uso
if __name__ == "__main__":
    # Configurar logging para prueba
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Generar reporte de ejemplo
    generator = ReportGenerator()
    excel_path = generator.generate_report()
    
    if excel_path:
        print(f"✅ Reporte Excel generado: {excel_path}")